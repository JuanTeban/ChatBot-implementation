import pandas as pd
import re
from openpyxl import load_workbook

EXCEL_FILE_PATH = "data/Seguimiento-hallazgos-Solman.xlsx"
SHEET_NAME = "Seguimiento "
HEADER_ROW = 0

def _sanitize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    new_columns = {}
    for col in df.columns:
        sanitized_col = re.sub(r'[^0-9a-zA-Z_]', '_', str(col).strip()).lower().strip('_')
        new_columns[col] = sanitized_col
    df = df.rename(columns=new_columns)
    return df

print("--- Cargando y preparando los datos del Excel... ---")
try:
    workbook = load_workbook(filename=EXCEL_FILE_PATH, data_only=True)
    sheet = workbook[SHEET_NAME]
    
    header_idx = HEADER_ROW + 1
    headers = [cell.value for cell in sheet[header_idx]]
    data_rows = [list(row) for row in sheet.iter_rows(min_row=header_idx + 1, values_only=True)]
    
    df = pd.DataFrame(data_rows, columns=headers)
    df = df.loc[:, df.columns.notna()]
    df = _sanitize_column_names(df)
    
    print("✅ Datos cargados y listos para la verificación.\n")

    print("#################################################################")
    print("### CATEGORÍA: PREGUNTAS DE RESUMEN Y AGRUPACIÓN ###")
    print("#################################################################\n")

    print("--- PREGUNTA: ¿Cómo vamos con los defectos? Dame un conteo por cada estado. ---")
    conteo_por_estado = df['estado_de_defecto'].value_counts()
    print("RESPUESTA CORRECTA:")
    print(conteo_por_estado.to_string(), "\n")

    print("--- PREGUNTA: ¿Qué módulo es el que tiene más hallazgos reportados? ---")
    modulo_con_mas_defectos = df['modulo'].value_counts().idxmax()
    conteo_maximo = df['modulo'].value_counts().max()
    print(f"RESPUESTA CORRECTA: El módulo con más defectos es '{modulo_con_mas_defectos}' con un total de {conteo_maximo}.\n")
    
    print("--- PREGUNTA: Hazme un resumen de quiénes son las personas que más defectos han reportado. ---")
    conteo_por_autor = df['autor_del_defecto'].value_counts().head(5)
    print("RESPUESTA CORRECTA: El top 5 de personas que más reportan es:")
    print(conteo_por_autor.to_string(), "\n")

    print("\n#################################################################")
    print("### CATEGORÍA: PREGUNTAS DE BÚSQUEDA ESPECÍFICA ###")
    print("#################################################################\n")

    print("--- PREGUNTA: Sobre el problema con la anulación de ajustes de EPM, ¿quién lo va a ver? ---")
    responsable_epm = df[df['defecto'].str.contains("ANULACIÓN AJUSTE", na=False)]['responsable_del_defecto']
    print("RESPUESTA CORRECTA:")
    print(responsable_epm.to_string(index=False), "\n")

    print("--- PREGUNTA: Busca el defecto que habla del 'comparativo por impuestos ICA', necesito saber quién lo reportó. ---")
    autor_ica = df[df['defecto'].str.contains("impuestos ICA", na=False)]['autor_del_defecto']
    print("RESPUESTA CORRECTA:")
    print(autor_ica.to_string(index=False), "\n")

    print("\n#################################################################")
    print("### CATEGORÍA: PREGUNTAS COMPLEJAS Y SITUACIONALES ###")
    print("#################################################################\n")

    print("--- PREGUNTA: Si soy del equipo técnico, ¿cuáles son los defectos bloqueantes que me deberían preocupar? ---")
    defectos_tecnicos_bloqueantes = df[(df['frente'] == 'TECNICO') & (df['bloqueante_escenarios'] == 'SI')]
    print("RESPUESTA CORRECTA:")
    print(defectos_tecnicos_bloqueantes[['n', 'defecto', 'responsable_del_defecto']].to_string(index=False), "\n")
    
    print("--- PREGUNTA: De los defectos que reportó MONICA PATRICIA, ¿cuántos siguen en estado nuevo? ---")
    defectos_monica_nuevos = df[
        (df['autor_del_defecto'].str.contains("MONICA PATRICIA", na=False)) & 
        (df['estado_de_defecto'] == 'Nuevo')
    ]
    conteo_monica = len(defectos_monica_nuevos)
    print(f"RESPUESTA CORRECTA: Hay {conteo_monica} defectos reportados por MONICA PATRICIA que siguen en estado 'Nuevo'.\n")

    print("--- PREGUNTA: Necesito un resumen de los defectos del módulo RE. ---")
    resumen_re = df[df['modulo'] == 'RE'][['n', 'defecto', 'estado_de_defecto', 'responsable_del_defecto']]
    print("RESPUESTA CORRECTA:")
    print(resumen_re.to_string(index=False))


except Exception as e:
    print(f"❌ Ocurrió un error: {e}")