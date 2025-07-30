import pandas as pd
import sqlite3
import re
import asyncio
from typing import Optional, Tuple, List, Set
from openpyxl import load_workbook

def _sanitize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    new_columns = {}
    for col in df.columns:
        sanitized_col = re.sub(r'[^0-9a-zA-Z_]', '_', str(col).strip()).lower().strip('_')
        new_columns[col] = sanitized_col
    df = df.rename(columns=new_columns)
    return df

async def create_db_from_excel(file_path: str) -> Optional[Tuple[sqlite3.Connection, str, List[str]]]:
    def blocking_task() -> Optional[Tuple[sqlite3.Connection, str, List[str]]]:
        try:
            print("Cargando datos desde Excel con openpyxl (data_only=True)...")
            workbook = load_workbook(filename=file_path, data_only=True)
            sheet = workbook["Seguimiento "]
            headers = [cell.value for cell in sheet[1]]
            data_rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.loc[:, df.columns.notna()]
            df_sanitized = _sanitize_column_names(df.copy())
            
            conn = sqlite3.connect(':memory:', check_same_thread=False)
            df_sanitized.to_sql("datos", conn, if_exists='replace', index=False)
            print(f"✅ Base de datos creada con {len(df_sanitized)} filas.")

            cursor = conn.cursor()
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='datos';")
            base_schema = cursor.fetchone()[0]
            
            enriched_schema_parts = [base_schema, "\n-- Valores de ejemplo para columnas categóricas:"]
            for column in df_sanitized.columns:
                if column.lower() in ['n', 'defecto', 'comentarios']: continue
                unique_values = df_sanitized[column].dropna().unique()
                if 1 < len(unique_values) <= 20:
                    values_str = ", ".join([f"'{str(v)}'" for v in unique_values])
                    enriched_schema_parts.append(f"-- Columna '{column}': [{values_str}]")
            enriched_schema = "\n".join(enriched_schema_parts)
            print("✅ Esquema enriquecido pre-calculado.")

            keywords: Set[str] = set()
            for col in df_sanitized.columns:
                words = col.split('_')
                keywords.update(word for word in words if len(word) > 2)
            keywords.update(['defecto', 'hallazgo', 'conteo', 'contar', 'cuántos', 'lista', 'dame', 'resumen', 'total', 'promedio'])
            keyword_list = sorted(list(keywords))
            print(f"✅ Palabras clave dinámicas generadas: {keyword_list}")
            
            return conn, enriched_schema, keyword_list

        except Exception as e:
            print(f"❌ ERROR al procesar el archivo Excel: {e}")
            return None
    
    result = await asyncio.to_thread(blocking_task)
    return result